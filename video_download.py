#!/usr/bin/env python
# coding: utf-8

# In[9]:


import time
import re
import openpyxl
# Open the Excel file
excel_file_path = 'video_dnld.xlsx'
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active

# Find the column index for "sno"
sno_column_index = None
for cell in sheet[1]:
    if cell.value == "Sno":
        sno_column_index = cell.column
        break

if sno_column_index is None:
    print("Column 'Sno' not found in the Excel sheet.")
else:
    # Loop through rows and extract data
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):  # Assuming URLs are in column D
        cell = row[0]
        video_url = cell.value
        sno = sheet.cell(row=cell.row, column=sno_column_index).value
        if video_url:
            try:
                yt = YouTube(video_url)
                stream = yt.streams.filter(file_extension='mp4').first()
                stream.download(filename=f'video_{sno}.mp4')
                # Write the extracted information back to the Excel file
                sheet.cell(row=cell.row, column=sno_column_index + 2).value = f'video_{sno}.mp4'
            except:
                print(f"Video {video_url} is unavailable. Skipping...")

    # Save the modified Excel file
    wb.save(excel_file_path)

