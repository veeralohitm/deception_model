#!/usr/bin/env python
# coding: utf-8

# In[34]:


from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from bs4 import BeautifulSoup
import time
import re
import openpyxl

# Set up the Chrome driver
chrome_service = ChromeService(executable_path='/Users/veeralohitm/Downloads/chromedriver-mac-arm64/chromedriver')
driver = webdriver.Chrome(service=chrome_service)

# Open the Excel file
excel_file_path = 'testdata copy.xlsx'
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active

# Find the column index for "sno"
sno_column_index = None
for cell in sheet[1]:
    if cell.value == "Sno":
        sno_column_index = cell.column
        break

if sno_column_index is None:
    print("Column 'Sno' not found in the Excel sheet.")
else:
    # Loop through rows and extract data
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):  # Assuming URLs are in column D
        cell = row[0]
        video_url = cell.value
        sno = sheet.cell(row=cell.row, column=sno_column_index).value
        if video_url:
            # Input YouTube video URL
            driver.get(video_url)

            # Wait for page to load (you might need to adjust the wait time)
            time.sleep(10)

            # Get the page source after the page is loaded
            page_source = driver.page_source

            # Parse the page source with BeautifulSoup
            soup = BeautifulSoup(page_source, 'html.parser')
            
            
            comment_element = soup.find('yt-formatted-string', class_='count-text')

            if comment_element:
            # Extract the text containing the comment count
                comment_text = comment_element.get_text(strip=True)
                # Extract the number from the text
                number_of_comments = ''.join(filter(str.isdigit, comment_text))
                
            else:
                number_of_comments = "N/A"
               
            # Extract Number of Likes 
            
            pattern = re.compile(r'like this video along with (\d{1,3}(,\d{3})*|\d+) other people')

            # Find the element containing the like count using the regular expression pattern
            element = soup.find(string=pattern)

            if element:
                match = pattern.search(element)
                number_of_likes = match.group(1).replace(',', '') if match else None
            else:
                number_of_likes = "N/A"
            
            # Extract Upload Date
            # Find all span elements
            span_elements = soup.find_all('span', class_='style-scope yt-formatted-string bold')

            # Check each span for a date pattern
            date = None
            for span in span_elements:
                text = span.get_text()
                match = re.search(r"(\w{3}\s\d{1,2},\s\d{4})", text)
                if match:
                    upload_date = match.group()
                    break
                else:
                    upload_date = "N/A"

            # Extract View Count
     
            view_count_element = soup.select_one('span.bold.style-scope.yt-formatted-string')
            view_count_text = view_count_element.get_text() if view_count_element else "N/A"

            # Remove " views" from the view count text
            view_count = view_count_text.replace(" views", "")
                  
            # Extract video title
            video_title = soup.find('h1', class_='style-scope ytd-watch-metadata')
            video_title = video_title.text.strip()if video_title else "N/A"
    
            # Extract number of likes
            likes_element = soup.find('yt-formatted-string', {'id': 'text', 'class': 'style-scope ytd-toggle-button-renderer style-text'})
            num_likes = likes_element.get_text() if likes_element else "N/A"

            # Extract channel name
            channel_name = soup.find('yt-formatted-string', {'class': 'ytd-channel-name'})
            channel_name=channel_name.text.strip()if channel_name else "N/A"
            
            # Extract number of subscribers
            subscriber_count_element = soup.find('yt-formatted-string', id='owner-sub-count')

            # Extract the text content of the element
            subscriber_count_text = subscriber_count_element.get_text()if subscriber_count_element else "N/A"

            # Extract number of subscribers
            subscriber_count_element = soup.find('yt-formatted-string', id='owner-sub-count')
            if subscriber_count_element:
            # Extract the text containing the subscriber count
                subscriber_text = subscriber_count_element.get_text(strip=True)
                # Extract the text before 'subscribers'
                subscriber_count = subscriber_text.split(' subscribers')[0]
            else:
                subscriber_count = "N/A"
        
            # Extract video description
            description_element = soup.find('span', class_='yt-core-attributed-string--link-inherit-color')
            video_description = description_element.get_text() if description_element else "N/A"

            # Extract top 100 comments
            comment_elements = soup.find_all('yt-formatted-string', {'class': 'style-scope ytd-comment-renderer'})
            comments = [comment.get_text() for comment in comment_elements]

            # Write the extracted information back to the Excel file
            sheet.cell(row=cell.row, column=sno_column_index + 4).value = video_title
            sheet.cell(row=cell.row, column=sno_column_index + 5).value = channel_name
            sheet.cell(row=cell.row, column=sno_column_index + 6).value = subscriber_count
            sheet.cell(row=cell.row, column=sno_column_index + 7).value = view_count
            sheet.cell(row=cell.row, column=sno_column_index + 8).value = number_of_likes
            sheet.cell(row=cell.row, column=sno_column_index + 9).value = upload_date
            sheet.cell(row=cell.row, column=sno_column_index + 10).value = number_of_comments
            sheet.cell(row=cell.row, column=sno_column_index + 11).value = video_description
            sheet.cell(row=cell.row, column=sno_column_index + 12).value = f'comments_{sno}.txt'

            # Save comments to a text file
            with open(f'comments_{sno}.txt', 'w', encoding='utf-8') as comment_file:
                for idx, comment in enumerate(comments[:100], start=1):
                    comment_file.write(f"{idx}: {comment}\n")

    # Save the modified Excel file
    wb.save(excel_file_path)

# Close the browser
driver.quit()

